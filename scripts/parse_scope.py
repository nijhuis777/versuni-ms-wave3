"""
Parse 'Mystery shopping 2026 - Inputs - updated all markets.xlsx'
and update config/scope.yaml with the real Wave III scope.

Run from the repo root:
    python scripts/parse_scope.py
"""

import sys
import json
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "pyyaml"])
    import openpyxl

import yaml

ROOT = Path(__file__).parent.parent
SCOPE_FILE = Path(r"C:\Users\MartijnNijhuis\dev\Versuni\2026\2026-001 Versuni Mystery Shopping - Wave III\Operations\Mystery shopping 2026 - Inputs - updated all markets.xlsx")
OUTPUT_YAML = ROOT / "config" / "scope.yaml"

MARKET_NAMES = {
    "DE": "Germany", "FR": "France", "NL": "Netherlands",
    "UK": "United Kingdom", "TR": "Turkey",
    "AU": "Australia", "BR": "Brazil", "US": "United States",
}

PLATFORM_MAP = {
    "DE": "roamler", "FR": "roamler", "NL": "roamler",
    "UK": "roamler", "TR": "roamler",
    "AU": "wiser", "US": "wiser", "BR": "pinion",
}


def print_sheet(ws):
    """Print all non-empty rows from a sheet."""
    for row in ws.iter_rows(values_only=True):
        if any(c is not None for c in row):
            print("  ", row)


def find_header_row(ws):
    """Find the row that looks like a header (contains 'market' or 'category' or similar)."""
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True), 1):
        vals = [str(v).lower() for v in row if v]
        if any(k in " ".join(vals) for k in ["market", "category", "country", "store"]):
            return i
    return 1


def parse_workbook():
    print(f"\nOpening: {SCOPE_FILE.name}")
    wb = openpyxl.load_workbook(SCOPE_FILE, data_only=True)
    print(f"Sheets found: {wb.sheetnames}\n")

    scope = {
        "markets": {},
        "categories": {
            "FAEM":           {"label": "Fully Automatic Espresso Machines", "wave2_json": "Versuni_FAEM.json"},
            "SAEM":           {"label": "Semi-Automatic Espresso Machines",  "wave2_json": None},
            "Airfryer":       {"label": "Airfryer",                          "wave2_json": None},
            "Blender":        {"label": "Blender",                           "wave2_json": None},
            "Iron":           {"label": "Iron (all types)",                  "wave2_json": None},
            "Handheld_Steamer": {"label": "Handheld Steamer",               "wave2_json": None},
            "Garlic_Press":   {"label": "Garlic Press",                      "wave2_json": None},
        },
        "kpis": {
            "KPI1": "Availability",
            "KPI2": "Visibility / Attractiveness",
            "KPI3": "Brand Recommendation",
        }
    }

    # Dump all sheets so we can see the structure
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"=== Sheet: {sheet_name} ({ws.max_row} rows × {ws.max_column} cols) ===")
        # Print first 40 rows
        for i, row in enumerate(ws.iter_rows(max_row=40, values_only=True), 1):
            if any(c is not None for c in row):
                # Truncate long values
                display = tuple(str(c)[:40] if c is not None else None for c in row)
                print(f"  row {i:2d}: {display}")
        print()

    # Save raw dump to file for inspection
    dump_path = ROOT / "scripts" / "scope_dump.txt"
    with open(dump_path, "w", encoding="utf-8") as f:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            f.write(f"\n{'='*60}\nSheet: {sheet_name}\n{'='*60}\n")
            for row in ws.iter_rows(values_only=True):
                if any(c is not None for c in row):
                    f.write(str(row) + "\n")

    print(f"\nFull dump saved → {dump_path}")
    print("Please share scope_dump.txt or paste the output above so the scope can be parsed correctly.")


if __name__ == "__main__":
    parse_workbook()
