"""
Questionnaire Export Tool — Multi-Platform
==========================================
Exports a questionnaire in the right format for each platform:
  - roamler  → JSON (upload directly to Roamler API)
  - wiser    → Excel (structured sheet for Wiser to configure their platform)
  - pinion   → Excel (structured sheet for Pinion)
  - review   → Human-readable Excel for internal review / client sign-off

Usage:
    python questionnaires/export_questionnaire.py --category FAEM --platform all
    python questionnaires/export_questionnaire.py --category FAEM --platform roamler
    python questionnaires/export_questionnaire.py --category FAEM --platform wiser
"""

import json
import argparse
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import yaml

ROOT = Path(__file__).parent.parent
OUTPUT_DIR = ROOT / "questionnaires" / "output"
EXPORT_DIR = ROOT / "questionnaires" / "exports"
EXPORT_DIR.mkdir(exist_ok=True)
CONFIG_DIR = ROOT / "config"

PLATFORM_COLORS = {
    "roamler": "4472C4",
    "wiser":   "70AD47",
    "pinion":  "ED7D31",
}

QUESTION_TYPE_LABELS = {
    1: "Single choice",
    2: "Multi choice",
    4: "Text / Number",
    5: "Photo",
    7: "Info / Instruction",
}


def load_config(name: str) -> dict:
    with open(CONFIG_DIR / f"{name}.yaml") as f:
        return yaml.safe_load(f)


def load_wave3_json(category: str) -> dict:
    path = OUTPUT_DIR / f"Versuni_{category}_Wave3.json"
    if not path.exists():
        raise FileNotFoundError(f"Wave III JSON not found at {path}. Run update_questionnaire.py first.")
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def flatten_questions(questionnaire: dict) -> list[dict]:
    """Extract questions into a flat list of dicts."""
    rows = []
    for q in questionnaire.get("Questions", []):
        answers = q.get("Answers", [])
        answer_texts = " | ".join(a.get("Text", "") for a in answers if a.get("Text"))
        condition = q.get("QuestionCondition")
        condition_str = ""
        if condition:
            conds = condition.get("Conditions", [])
            condition_str = " AND ".join(
                f"{c.get('QuestionCode')} = {c.get('AnswerCode')}" for c in conds
            )
        rows.append({
            "Seq": q.get("Sequence", ""),
            "Code": q.get("Code", ""),
            "Type": QUESTION_TYPE_LABELS.get(q.get("Type"), str(q.get("Type", ""))),
            "Text": q.get("Text", ""),
            "Optional": "Yes" if q.get("IsOptional") else "No",
            "Answers": answer_texts,
            "Condition": condition_str,
            "KPI": "KPI" in q.get("Code", ""),
        })
    return rows


def export_review_excel(category: str, questionnaire: dict, platform: str = "review"):
    rows = flatten_questions(questionnaire)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{category} Questionnaire"

    headers = ["Seq", "Code", "Type", "Question Text", "Optional", "Answer Options", "Show If", "KPI?"]
    header_fill = PatternFill("solid", fgColor=PLATFORM_COLORS.get(platform, "404040"))
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    kpi_fill = PatternFill("solid", fgColor="FFF2CC")
    for row_idx, row in enumerate(rows, 2):
        values = [
            row["Seq"], row["Code"], row["Type"], row["Text"],
            row["Optional"], row["Answers"], row["Condition"],
            "✓" if row["KPI"] else ""
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row["KPI"]:
                cell.fill = kpi_fill

    # Column widths
    widths = [6, 30, 16, 60, 10, 60, 40, 8]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(rows)+1}"

    filename = EXPORT_DIR / f"Versuni_{category}_Wave3_{platform}.xlsx"
    wb.save(filename)
    print(f"  Exported ({platform}) → {filename}")
    return filename


def export_roamler_json(category: str, questionnaire: dict):
    path = EXPORT_DIR / f"Versuni_{category}_Wave3_roamler.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(questionnaire, f, indent=2, ensure_ascii=False)
    print(f"  Exported (roamler) → {path}")


def main():
    parser = argparse.ArgumentParser(description="Export questionnaire for platform")
    parser.add_argument("--category", "-c", required=True)
    parser.add_argument("--platform", "-p", default="all",
                        choices=["all", "roamler", "wiser", "pinion", "review"])
    args = parser.parse_args()

    questionnaire = load_wave3_json(args.category)
    scope = load_config("scope")
    cat_label = scope["categories"].get(args.category, {}).get("label", args.category)

    print(f"\nExporting: {cat_label} (Wave III)")

    platforms = ["roamler", "wiser", "pinion", "review"] if args.platform == "all" else [args.platform]

    for platform in platforms:
        if platform == "roamler":
            export_roamler_json(args.category, questionnaire)
        else:
            # Wiser and Pinion get the structured Excel review format
            export_review_excel(args.category, questionnaire, platform)


if __name__ == "__main__":
    main()
